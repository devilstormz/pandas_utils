import pandas as pd
from datetime import datetime
from typing import Dict, List, Tuple, Optional
from dataclasses import dataclass
from enum import Enum
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


class TradeType(Enum):
    REPO = "REPO"
    REVERSE_REPO = "REVERSE_REPO"
    BOND_BORROW = "BOND_BORROW"
    BOND_LEND = "BOND_LEND"
    COLLATERAL_SWAP = "COLLATERAL_SWAP"
    TRI_PARTY_REPO = "TRI_PARTY_REPO"
    TRI_PARTY_REVERSE = "TRI_PARTY_REVERSE"


class FlowDirection(Enum):
    SOURCE = "SOURCE"  # Cash in / Collateral in
    USE = "USE"       # Cash out / Collateral out


@dataclass
class Trade:
    trade_id: str
    trade_type: TradeType
    isin: str
    currency: str
    nominal_amount: float
    cash_amount: float
    rate: float
    start_date: datetime
    maturity_date: datetime
    counterparty: str
    trader: str
    
    
@dataclass
class CollateralSwap:
    swap_id: str
    repo_trade: Trade
    reverse_repo_trade: Trade
    net_cash_flow: float
    

@dataclass
class FlowMapping:
    isin: str
    source_trade: Trade
    use_trades: List[Trade]
    unused_amount: float
    

class SourcesUsesClassifier:
    """Classifies trades into sources and uses based on trade type"""
    
    @staticmethod
    def classify_trade(trade: Trade) -> Tuple[FlowDirection, str]:
        """
        Returns (FlowDirection, description) for a trade
        """
        classification_map = {
            TradeType.REPO: (FlowDirection.SOURCE, "Cash received via repo"),
            TradeType.REVERSE_REPO: (FlowDirection.USE, "Cash paid via reverse repo"), 
            TradeType.BOND_BORROW: (FlowDirection.SOURCE, "Collateral received via borrow"),
            TradeType.BOND_LEND: (FlowDirection.USE, "Collateral lent out"),
            TradeType.TRI_PARTY_REPO: (FlowDirection.SOURCE, "Cash received via tri-party repo"),
            TradeType.TRI_PARTY_REVERSE: (FlowDirection.USE, "Cash paid via tri-party reverse"),
            TradeType.COLLATERAL_SWAP: (FlowDirection.SOURCE, "Net cash via collateral swap")
        }
        
        return classification_map.get(trade.trade_type, (FlowDirection.USE, "Unknown"))
    
    @staticmethod
    def get_cash_sources(trades: List[Trade]) -> List[Trade]:
        """Returns trades that are cash sources"""
        return [t for t in trades if SourcesUsesClassifier.classify_trade(t)[0] == FlowDirection.SOURCE]
    
    @staticmethod
    def get_cash_uses(trades: List[Trade]) -> List[Trade]:
        """Returns trades that are cash uses"""  
        return [t for t in trades if SourcesUsesClassifier.classify_trade(t)[0] == FlowDirection.USE]


class CollateralSwapIdentifier:
    """Identifies and creates collateral swap structures from repo/reverse repo pairs"""
    
    def __init__(self, tolerance_days: int = 2, tolerance_amount: float = 0.01):
        self.tolerance_days = tolerance_days
        self.tolerance_amount = tolerance_amount
    
    def identify_swaps(self, trades: List[Trade]) -> List[CollateralSwap]:
        """
        Identifies collateral swaps by matching repo/reverse repo pairs
        """
        swaps = []
        repos = [t for t in trades if t.trade_type == TradeType.REPO]
        reverse_repos = [t for t in trades if t.trade_type == TradeType.REVERSE_REPO]
        
        used_reverse_repos = set()
        
        for repo in repos:
            for reverse_repo in reverse_repos:
                if (reverse_repo.trade_id in used_reverse_repos or 
                    not self._is_potential_swap(repo, reverse_repo)):
                    continue
                
                swap = CollateralSwap(
                    swap_id=f"SWAP_{repo.trade_id}_{reverse_repo.trade_id}",
                    repo_trade=repo,
                    reverse_repo_trade=reverse_repo,
                    net_cash_flow=repo.cash_amount - reverse_repo.cash_amount
                )
                swaps.append(swap)
                used_reverse_repos.add(reverse_repo.trade_id)
                break
        
        return swaps
    
    def _is_potential_swap(self, repo: Trade, reverse_repo: Trade) -> bool:
        """Check if repo and reverse repo could form a collateral swap"""
        # Same ISIN
        if repo.isin != reverse_repo.isin:
            return False
            
        # Similar amounts (within tolerance)
        amount_diff = abs(repo.nominal_amount - reverse_repo.nominal_amount)
        if amount_diff > self.tolerance_amount * repo.nominal_amount:
            return False
            
        # Similar dates (within tolerance)
        date_diff = abs((repo.start_date - reverse_repo.start_date).days)
        if date_diff > self.tolerance_days:
            return False
            
        return True


class FlowTraceabilityEngine:
    """Builds flow traceability mapping for ISINs"""
    
    def __init__(self):
        self.classifier = SourcesUsesClassifier()
    
    def build_flow_mapping(self, trades: List[Trade]) -> List[FlowMapping]:
        """
        Creates traceability mapping showing how each ISIN flows from source to use
        """
        # Group trades by ISIN
        isin_groups = {}
        for trade in trades:
            if trade.isin not in isin_groups:
                isin_groups[trade.isin] = []
            isin_groups[trade.isin].append(trade)
        
        flow_mappings = []
        
        for isin, isin_trades in isin_groups.items():
            # Separate sources and uses for this ISIN
            sources = [t for t in isin_trades if self._is_source_trade(t)]
            uses = [t for t in isin_trades if self._is_use_trade(t)]
            
            # Create mappings for each source
            for source in sources:
                # Find matching uses (simplified allocation logic)
                allocated_uses = self._allocate_uses_to_source(source, uses)
                
                unused_amount = source.nominal_amount - sum(u.nominal_amount for u in allocated_uses)
                
                flow_mapping = FlowMapping(
                    isin=isin,
                    source_trade=source,
                    use_trades=allocated_uses,
                    unused_amount=max(0, unused_amount)
                )
                flow_mappings.append(flow_mapping)
        
        return flow_mappings
    
    def _is_source_trade(self, trade: Trade) -> bool:
        """Check if trade is a source of collateral"""
        return trade.trade_type in [TradeType.REVERSE_REPO, TradeType.BOND_BORROW, TradeType.TRI_PARTY_REVERSE]
    
    def _is_use_trade(self, trade: Trade) -> bool:
        """Check if trade is a use of collateral"""
        return trade.trade_type in [TradeType.REPO, TradeType.BOND_LEND, TradeType.TRI_PARTY_REPO]
    
    def _allocate_uses_to_source(self, source: Trade, available_uses: List[Trade]) -> List[Trade]:
        """
        Simplified allocation logic - could be enhanced with more sophisticated matching
        """
        allocated = []
        remaining_amount = source.nominal_amount
        
        # Sort uses by date to prioritize earlier deployments
        sorted_uses = sorted(available_uses, key=lambda x: x.start_date)
        
        for use in sorted_uses:
            if remaining_amount <= 0:
                break
                
            if use.nominal_amount <= remaining_amount:
                allocated.append(use)
                remaining_amount -= use.nominal_amount
        
        return allocated


class SourcesUsesReporter:
    """Generates Excel reports for sources and uses analysis"""
    
    def __init__(self):
        self.classifier = SourcesUsesClassifier()
        self.swap_identifier = CollateralSwapIdentifier()
        self.flow_engine = FlowTraceabilityEngine()
    
    def generate_report(self, trades: List[Trade], output_file: str):
        """Generate comprehensive sources and uses Excel report"""
        
        # Identify collateral swaps
        swaps = self.swap_identifier.identify_swaps(trades)
        
        # Build flow mappings
        flow_mappings = self.flow_engine.build_flow_mapping(trades)
        
        # Create Excel workbook
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Generate individual sheets
        self._create_net_flow_summary(wb, trades)
        self._create_traceability_mapping(wb, flow_mappings)
        self._create_unused_collateral_report(wb, flow_mappings)
        self._create_collateral_swaps_report(wb, swaps)
        self._create_trade_detail_sheet(wb, trades)
        
        # Save workbook
        wb.save(output_file)
        print(f"Sources and Uses report generated: {output_file}")
    
    def _create_net_flow_summary(self, wb, trades: List[Trade]):
        """Create net flow summary by currency"""
        ws = wb.create_sheet("Net Flow Summary")
        
        # Headers
        headers = ["Currency", "Cash Sources", "Cash Uses", "Net Position", "Source Details", "Use Details"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        # Group by currency
        currency_flows = {}
        for trade in trades:
            if trade.currency not in currency_flows:
                currency_flows[trade.currency] = {"sources": [], "uses": []}
            
            flow_direction, description = self.classifier.classify_trade(trade)
            if flow_direction == FlowDirection.SOURCE:
                currency_flows[trade.currency]["sources"].append(trade)
            else:
                currency_flows[trade.currency]["uses"].append(trade)
        
        # Populate data
        row = 2
        for currency, flows in currency_flows.items():
            total_sources = sum(t.cash_amount for t in flows["sources"])
            total_uses = sum(t.cash_amount for t in flows["uses"])
            net_position = total_sources - total_uses
            
            source_details = "; ".join([f"{t.trade_type.value}: {t.cash_amount:,.0f}" for t in flows["sources"][:3]])
            use_details = "; ".join([f"{t.trade_type.value}: {t.cash_amount:,.0f}" for t in flows["uses"][:3]])
            
            ws.cell(row=row, column=1, value=currency)
            ws.cell(row=row, column=2, value=total_sources)
            ws.cell(row=row, column=3, value=total_uses)
            ws.cell(row=row, column=4, value=net_position)
            ws.cell(row=row, column=5, value=source_details)
            ws.cell(row=row, column=6, value=use_details)
            
            # Color code net position
            net_cell = ws.cell(row=row, column=4)
            if net_position > 0:
                net_cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            elif net_position < 0:
                net_cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
            
            row += 1
        
        # Auto-fit columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_traceability_mapping(self, wb, flow_mappings: List[FlowMapping]):
        """Create traceability mapping sheet"""
        ws = wb.create_sheet("Traceability Mapping")
        
        headers = ["ISIN", "Source Trade ID", "Source Type", "Source Amount", "Source Date", 
                  "Use Trade IDs", "Use Types", "Use Amounts", "Total Used", "Unused Amount"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        row = 2
        for mapping in flow_mappings:
            use_trade_ids = "; ".join([t.trade_id for t in mapping.use_trades])
            use_types = "; ".join([t.trade_type.value for t in mapping.use_trades])
            use_amounts = "; ".join([f"{t.nominal_amount:,.0f}" for t in mapping.use_trades])
            total_used = sum(t.nominal_amount for t in mapping.use_trades)
            
            ws.cell(row=row, column=1, value=mapping.isin)
            ws.cell(row=row, column=2, value=mapping.source_trade.trade_id)
            ws.cell(row=row, column=3, value=mapping.source_trade.trade_type.value)
            ws.cell(row=row, column=4, value=mapping.source_trade.nominal_amount)
            ws.cell(row=row, column=5, value=mapping.source_trade.start_date.strftime("%Y-%m-%d"))
            ws.cell(row=row, column=6, value=use_trade_ids)
            ws.cell(row=row, column=7, value=use_types)
            ws.cell(row=row, column=8, value=use_amounts)
            ws.cell(row=row, column=9, value=total_used)
            ws.cell(row=row, column=10, value=mapping.unused_amount)
            
            # Highlight unused collateral
            if mapping.unused_amount > 0:
                unused_cell = ws.cell(row=row, column=10)
                unused_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            
            row += 1
        
        # Auto-fit columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_unused_collateral_report(self, wb, flow_mappings: List[FlowMapping]):
        """Create unused collateral identification sheet"""
        ws = wb.create_sheet("Unused Collateral")
        
        unused_mappings = [m for m in flow_mappings if m.unused_amount > 0]
        
        headers = ["ISIN", "Source Trade ID", "Source Type", "Counterparty", "Total Amount", 
                  "Used Amount", "Unused Amount", "Days Outstanding", "Risk Flag"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        row = 2
        for mapping in unused_mappings:
            used_amount = sum(t.nominal_amount for t in mapping.use_trades)
            days_outstanding = (datetime.now() - mapping.source_trade.start_date).days
            
            # Risk flag based on days outstanding
            risk_flag = "HIGH" if days_outstanding > 30 else "MEDIUM" if days_outstanding > 7 else "LOW"
            
            ws.cell(row=row, column=1, value=mapping.isin)
            ws.cell(row=row, column=2, value=mapping.source_trade.trade_id)
            ws.cell(row=row, column=3, value=mapping.source_trade.trade_type.value)
            ws.cell(row=row, column=4, value=mapping.source_trade.counterparty)
            ws.cell(row=row, column=5, value=mapping.source_trade.nominal_amount)
            ws.cell(row=row, column=6, value=used_amount)
            ws.cell(row=row, column=7, value=mapping.unused_amount)
            ws.cell(row=row, column=8, value=days_outstanding)
            ws.cell(row=row, column=9, value=risk_flag)
            
            # Color code risk
            risk_cell = ws.cell(row=row, column=9)
            if risk_flag == "HIGH":
                risk_cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                risk_cell.font = Font(color="FFFFFF", bold=True)
            elif risk_flag == "MEDIUM":
                risk_cell.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
            
            row += 1
        
        # Auto-fit columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 25)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_collateral_swaps_report(self, wb, swaps: List[CollateralSwap]):
        """Create collateral swaps identification sheet"""
        ws = wb.create_sheet("Collateral Swaps")
        
        headers = ["Swap ID", "ISIN", "Repo Trade ID", "Repo Amount", "Reverse Trade ID", 
                  "Reverse Amount", "Net Cash Flow", "Repo Rate", "Reverse Rate", "Rate Spread"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4ECDC4", end_color="4ECDC4", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        row = 2
        for swap in swaps:
            rate_spread = swap.repo_trade.rate - swap.reverse_repo_trade.rate
            
            ws.cell(row=row, column=1, value=swap.swap_id)
            ws.cell(row=row, column=2, value=swap.repo_trade.isin)
            ws.cell(row=row, column=3, value=swap.repo_trade.trade_id)
            ws.cell(row=row, column=4, value=swap.repo_trade.cash_amount)
            ws.cell(row=row, column=5, value=swap.reverse_repo_trade.trade_id)
            ws.cell(row=row, column=6, value=swap.reverse_repo_trade.cash_amount)
            ws.cell(row=row, column=7, value=swap.net_cash_flow)
            ws.cell(row=row, column=8, value=f"{swap.repo_trade.rate:.4f}%")
            ws.cell(row=row, column=9, value=f"{swap.reverse_repo_trade.rate:.4f}%")
            ws.cell(row=row, column=10, value=f"{rate_spread:.4f}%")
            
            # Highlight profitable swaps
            if swap.net_cash_flow > 0:
                net_cell = ws.cell(row=row, column=7)
                net_cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            
            row += 1
        
        # Auto-fit columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_trade_detail_sheet(self, wb, trades: List[Trade]):
        """Create detailed trade listing sheet"""
        ws = wb.create_sheet("Trade Details")
        
        headers = ["Trade ID", "Trade Type", "ISIN", "Currency", "Nominal Amount", 
                  "Cash Amount", "Rate", "Start Date", "Maturity Date", "Counterparty", "Trader"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="95A5A6", end_color="95A5A6", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        row = 2
        for trade in trades:
            ws.cell(row=row, column=1, value=trade.trade_id)
            ws.cell(row=row, column=2, value=trade.trade_type.value)
            ws.cell(row=row, column=3, value=trade.isin)
            ws.cell(row=row, column=4, value=trade.currency)
            ws.cell(row=row, column=5, value=trade.nominal_amount)
            ws.cell(row=row, column=6, value=trade.cash_amount)
            ws.cell(row=row, column=7, value=f"{trade.rate:.4f}%")
            ws.cell(row=row, column=8, value=trade.start_date.strftime("%Y-%m-%d"))
            ws.cell(row=row, column=9, value=trade.maturity_date.strftime("%Y-%m-%d"))
            ws.cell(row=row, column=10, value=trade.counterparty)
            ws.cell(row=row, column=11, value=trade.trader)
            
            row += 1
        
        # Auto-fit columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 25)
            ws.column_dimensions[column_letter].width = adjusted_width


# Example usage and testing
def create_sample_data() -> List[Trade]:
    """Create sample trade data for testing"""
    from datetime import timedelta
    
    base_date = datetime(2024, 1, 15)
    
    sample_trades = [
        # USD Repos and Reverse Repos
        Trade("RPO001", TradeType.REPO, "US912828Z293", "USD", 10000000, 10000000, 5.25, 
              base_date, base_date + timedelta(days=30), "GOLDMAN SACHS", "TRADER_A"),
        Trade("REV001", TradeType.REVERSE_REPO, "US912828Z293", "USD", 10000000, 10000000, 5.15, 
              base_date, base_date + timedelta(days=30), "MORGAN STANLEY", "TRADER_A"),
        
        # Bond Borrow/Lend
        Trade("BBW001", TradeType.BOND_BORROW, "US037833100", "USD", 5000000, 5000000, 0.25, 
              base_date, base_date + timedelta(days=14), "DEUTSCHE BANK", "TRADER_B"),
        Trade("BLN001", TradeType.BOND_LEND, "US037833100", "USD", 3000000, 3000000, 0.35, 
              base_date + timedelta(days=1), base_date + timedelta(days=14), "BARCLAYS", "TRADER_B"),
        
        # EUR Tri-party
        Trade("TPR001", TradeType.TRI_PARTY_REPO, "DE0001102309", "EUR", 8000000, 8000000, 3.75, 
              base_date, base_date + timedelta(days=7), "BNP PARIBAS", "TRADER_C"),
        Trade("TPV001", TradeType.TRI_PARTY_REVERSE, "DE0001102309", "EUR", 8000000, 8000000, 3.65, 
              base_date, base_date + timedelta(days=7), "SOCIETE GENERALE", "TRADER_C"),
        
        # Additional trades for more complex scenarios
        Trade("RPO002", TradeType.REPO, "GB00B24FF097", "GBP", 6000000, 6000000, 4.85, 
              base_date + timedelta(days=2), base_date + timedelta(days=21), "HSBC", "TRADER_D"),
        Trade("REV002", TradeType.REVERSE_REPO, "GB00B24FF097", "GBP", 8000000, 8000000, 4.75, 
              base_date, base_date + timedelta(days=21), "LLOYDS", "TRADER_D"),
    ]
    
    return sample_trades


def main():
    """Main function to demonstrate the reporting system"""
    
    # Create sample data
    trades = create_sample_data()
    
    # Initialize reporter
    reporter = SourcesUsesReporter()
    
    # Generate report
    output_file = "sources_uses_report.xlsx"
    reporter.generate_report(trades, output_file)
    
    # Print summary statistics
    print("\n=== SOURCES AND USES SUMMARY ===")
    
    sources = reporter.classifier.get_cash_sources(trades)
    uses = reporter.classifier.get_cash_uses(trades)
    
    print(f"Total Cash Sources: ${sum(t.cash_amount for t in sources):,.2f}")
    print(f"Total Cash Uses: ${sum(t.cash_amount for t in uses):,.2f}")
    print(f"Net Position: ${sum(t.cash_amount for t in sources) - sum(t.cash_amount for t in uses):,.2f}")
    
    # Identify swaps
    swaps = reporter.swap_identifier.identify_swaps(trades)
    print(f"\nIdentified {len(swaps)} collateral swaps")
    
    # Build flow mappings
    flow_mappings = reporter.flow_engine.build_flow_mapping(trades)
    unused_collateral = [m for m in flow_mappings if m.unused_amount > 0]
    print(f"Found {len(unused_collateral)} positions with unused collateral")


if __name__ == "__main__":
    main()
